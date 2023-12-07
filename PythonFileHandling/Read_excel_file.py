"""
pip install openpyxl
"""
import openpyxl
"""
filename = "test_data.xlsx"

wb_obj = openpyxl.load_workbook(filename)

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row = 1, column = 1)
print(cell_obj.value)
"""
def read_cell_data(filename, row_num, col_num):
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = row_num, column = col_num)
    return cell_obj.value

#var1 = read_cell_data('test_data.xlsx', 1, 1)
#print(var1) # Pune
#var2 = read_cell_data('test_data.xlsx', 2, 1)
#print(var2) # Mumbai


# print first column values
# for row_num in range(1, 7):
#     val = read_cell_data('test_data.xlsx', row_num, 1)
#     print(val)


#read all data from row and colum
"""
for row_num in range(1, 7):
    for col_num in range(1, 4):
        val = read_cell_data('test_data.xlsx', row_num, col_num)
        print(val, end=" | ")
    print()
"""

def get_all_rows_col_values(filename):
    wb_obj = openpyxl.load_workbook(filename)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    for i in range(1, max_row+1):
        for j in range(1, max_col+1):
            cell_obj = sheet_obj.cell(row= i,  column = j)
            print(cell_obj.value, end=" | ")
        print()

get_all_rows_col_values("test_data.xlsx")